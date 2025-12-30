import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image


# ---------------- THEME ----------------
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class CashManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cash Manager")
        self.root.geometry("450x450")
        self.root.configure(fg_color="#0f172a")  # dark blue

        self.transaction_file = "transactions.xlsx"
        self.categories_file = "categories.xlsx"

        self.initialize_transaction_file()
        self.initialize_categories_file()

        self.available_cash = ctk.StringVar()
        self.update_available_cash()

        self.create_ui()

    # ---------------- UI ----------------
    def create_ui(self):
        # Top box
        top_box = ctk.CTkFrame(self.root, height=80, fg_color="#020617")
        top_box.pack(fill="x", padx=20, pady=10)

        ctk.CTkLabel(
            top_box,
            text="Cash Manager",
            font=("Segoe UI", 28, "bold"),
            text_color="#e5e7eb"
        ).pack(pady=5)

        ctk.CTkLabel(
            top_box,
            textvariable=self.available_cash,
            font=("Segoe UI", 16),
            text_color="#22c55e"
        ).pack(pady=10)
        ctk.CTkButton(
            top_box,
            text="↻ Refresh",
            width=80,
            height=28,
            fg_color="#1e293b",
            hover_color="#334155",
            font=("Segoe UI", 12),
            command=self.refresh_cash_from_excel
        ).pack(pady=5)

        # Tabs
        tab_frame = ctk.CTkFrame(self.root, fg_color="#020617")
        tab_frame.pack(fill="both", expand=True, padx=20, pady=10)

        tabs = ctk.CTkTabview(tab_frame, fg_color="#020617")
        tabs.pack(fill="both", expand=True, padx=20, pady=10)
        tabs._segmented_button.configure(
            height=42,
            font=("Segoe UI", 14, "bold"),
            corner_radius=18
        )

        tab_cash = tabs.add("Cash")
        tab_statements = tabs.add("Statements")
        tab_analysis = tabs.add("Analysis")

        # ---- CASH TAB ----
        cash_button_frame = ctk.CTkFrame(tab_cash, fg_color="transparent")
        cash_button_frame.pack(pady=40)

        add_img = Image.open("icons/add.png")

        add_icon = ctk.CTkImage(
            light_image=add_img,
            dark_image=add_img,
            size=(64, 64)
        )

        remove_img = Image.open("icons/delete.png")

        remove_icon = ctk.CTkImage(
            light_image=remove_img,
            dark_image=remove_img,
            size=(64, 64)
        )


        add_cash_btn = ctk.CTkButton(
            cash_button_frame,
            text="Add Cash",
            image=add_icon,
            compound="top",
            width=125,
            height=100,
            fg_color="#0f172a",
            hover_color="#15803d",
            font=("Segoe UI", 16, "bold"),
            command=self.add_cash_popup,corner_radius=10
        )
        add_cash_btn.grid(row=0, column=0, padx=10)

        remove_cash_btn = ctk.CTkButton(
            cash_button_frame,
            text="Remove Cash",
            image=remove_icon,
            compound="top",
            width=70,
            height=70,
            fg_color="#0f172a",
            hover_color="#b91c1c",
            font=("Segoe UI", 16, "bold"),
            command=self.remove_cash_popup,
            corner_radius=10
        )
        remove_cash_btn.grid(row=0, column=1, padx=30)

        # ---- STATEMENTS TAB ----
        statement_frame = ctk.CTkFrame(tab_statements, fg_color="transparent")
        statement_frame.pack(pady=40)

        monthly_img = Image.open("icons/monthly.png")
        range_img = Image.open("icons/date.png")
        all_img = Image.open("icons/statement.png")

        monthly_icon = ctk.CTkImage(
            light_image=monthly_img,
            dark_image=monthly_img,
            size=(40, 40)
        )

        range_icon = ctk.CTkImage(
            light_image=range_img,
            dark_image=range_img,
            size=(40, 40)
        )

        all_icon = ctk.CTkImage(
            light_image=all_img,
            dark_image=all_img,
            size=(40, 40)
        )

        monthly_btn = ctk.CTkButton(
            statement_frame,
            text="Monthly\nStatement",
            image=monthly_icon,
            compound="top",
            width=100,
            height=100,
            fg_color="#0f172a",
            hover_color="#15803d",
            font=("Serif Sans", 13,"bold"),
            command=self.show_monthly_statement
        )
        monthly_btn.grid(row=0, column=0, padx=10)

        range_btn = ctk.CTkButton(
            statement_frame,
            text="Date Range\nStatement",
            image=range_icon,
            compound="top",
            width=100,
            height=100,
            fg_color="#0f172a",
            hover_color="#4A1470",
            font=("Serif Sans", 13,"bold"),
            command=self.show_date_range_statement
        )
        range_btn.grid(row=0, column=1, padx=10)

        all_btn = ctk.CTkButton(
            statement_frame,
            text="View\nStatement",
            image=all_icon,
            compound="top",
            width=100,
            height=100,
            fg_color="#0f172a",
            hover_color="#b91c1c",
            font=("Serif Sans", 13, "bold"),
            command=self.view_statement
        )
        all_btn.grid(row=0, column=2, padx=10)

        # ---- ANALYSIS TAB ----
        manage_button_frame = ctk.CTkFrame(tab_analysis, fg_color="transparent")
        manage_button_frame.pack(pady=40)
        cat_img = Image.open("icons/categories.png")

        cat_icon = ctk.CTkImage(
            light_image=cat_img,
            dark_image=cat_img,
            size=(64, 64)
        )

        pie_img = Image.open("icons/pie.png")

        pie_icon = ctk.CTkImage(
            light_image=pie_img,
            dark_image=pie_img,
            size=(64, 64)
        )

        cat_btn = ctk.CTkButton(
            manage_button_frame,
            text="Manage Categories",
            image=cat_icon,
            compound="top",
            width=125,
            height=100,
            fg_color="#0f172a",
            hover_color="#15803d",
            font=("Segoe UI", 12, "bold"),
            command=self.manage_categories, corner_radius=10
        )
        cat_btn.grid(row=0, column=0, padx=10)

        pie_btn = ctk.CTkButton(
            manage_button_frame,
            text="Analysis",
            image=pie_icon,
            compound="top",
            width=120,
            height=70,
            fg_color="#0f172a",
            hover_color="#b91c1c",
            font=("Segoe UI", 16, "bold"),
            command=self.show_analysis,
            corner_radius=10
        )
        pie_btn.grid(row=0, column=1, padx=30)

    # ---------------- UTIL ----------------
    def center_popup(self, popup, w, h):
        self.root.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        popup.geometry(f"{w}x{h}+{x}+{y}")

    # ---------------- FILE INIT ----------------
    def initialize_transaction_file(self):
        if not os.path.exists(self.transaction_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Type", "Date", "Amount", "Category"])
            wb.save(self.transaction_file)

    def initialize_categories_file(self):
        if not os.path.exists(self.categories_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Category"])
            ws.append(["General"])
            wb.save(self.categories_file)

    # ---------------- CASH ----------------
    def get_available_cash(self):
        df = pd.read_excel(self.transaction_file)
        return df[df["Type"] == "Add"]["Amount"].sum() - df[df["Type"] == "Remove"]["Amount"].sum()

    def update_available_cash(self):
        self.available_cash.set(f"Available Cash: ₹{self.get_available_cash():.2f}")

    def record_transaction(self, ttype, amount, category):
        wb = load_workbook(self.transaction_file)
        ws = wb.active
        ws.append([ttype, datetime.now().strftime("%d-%m-%Y"), amount, category])
        wb.save(self.transaction_file)

    def add_cash_popup(self):
        self.manage_cash_popup("Add Cash", "Add")

    def remove_cash_popup(self):
        self.manage_cash_popup("Remove Cash", "Remove")

    def manage_cash_popup(self, title, ttype):
        popup = ctk.CTkToplevel(self.root, fg_color="gray14")
        popup.title(title)
        popup.grab_set()
        popup.lift()
        self.center_popup(popup, 350, 250)

        amount_var = ctk.StringVar()
        category_var = ctk.StringVar(value=self.get_categories()[0])

        ctk.CTkLabel(popup, text="Amount").pack(pady=5)
        ctk.CTkEntry(popup, textvariable=amount_var).pack()

        ctk.CTkLabel(popup, text="Category").pack(pady=5)
        ctk.CTkOptionMenu(popup, variable=category_var, values=self.get_categories()).pack()

        def submit():
            try:
                amt = float(amount_var.get())
                self.record_transaction(ttype, amt, category_var.get())
                self.update_available_cash()
                popup.destroy()
            except:
                messagebox.showerror("Error", "Invalid amount")

        ctk.CTkButton(popup, text="Submit", command=submit).pack(pady=15)

    # ---------------- CATEGORIES ----------------
    def get_categories(self):
        return pd.read_excel(self.categories_file)["Category"].tolist()

    def manage_categories(self):

        popup = ctk.CTkToplevel(self.root)
        popup.title("Manage Categories")
        popup.geometry("320x360")
        popup.configure(fg_color="gray14")
        popup.transient(self.root)
        popup.grab_set()

        # ---------- Canvas ----------
        canvas = ctk.CTkCanvas(
            popup,
            bg="gray14",
            highlightthickness=0
        )

        canvas.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=10)

        scrollbar = ctk.CTkScrollbar(popup, command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        canvas.configure(yscrollcommand=scrollbar.set)

        categories_frame = ctk.CTkFrame(canvas, fg_color="gray17")
        canvas.create_window(
            (0, 0),
            window=categories_frame,
            anchor="nw",
            width=280
        )

        # ---------- FUNCTIONS ----------
        def refresh():
            for widget in categories_frame.winfo_children():
                widget.destroy()

            categories = self.get_categories()
            for category in categories:
                row = ctk.CTkFrame(categories_frame, fg_color="transparent")
                row.pack(fill="x", pady=4, padx=5)

                ctk.CTkLabel(
                    row,
                    text=category,
                    anchor="w",
                    font=("Serif Sans", 12),
                    text_color="#DCE4EE"
                ).pack(side="left", fill="x", expand=True)

                ctk.CTkButton(
                    row,
                    text="✖",
                    width=30,
                    fg_color="#8B0000",
                    hover_color="#5C0000",
                    command=lambda c=category: delete_category(c)
                ).pack(side="right")

            categories_frame.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        def add_category():
            dialog = ctk.CTkInputDialog(
                title="Add Category",
                text="Enter new category:"
            )
            new_category = dialog.get_input()

            if not new_category:
                return

            try:
                wb = load_workbook(self.categories_file)
                sheet = wb.active
                sheet.append([new_category.strip()])
                wb.save(self.categories_file)
                refresh()
            except Exception as e:
                messagebox.showerror("Error", str(e))

        def delete_category(category):
            if not messagebox.askyesno(
                    "Confirm Delete",
                    f"Delete category '{category}'?"
            ):
                return

            wb = load_workbook(self.categories_file)
            sheet = wb.active

            for row in range(1, sheet.max_row + 1):
                if sheet.cell(row, 1).value == category:
                    sheet.delete_rows(row)
                    break

            wb.save(self.categories_file)
            refresh()

        # ---------- Bottom Button ----------
        ctk.CTkButton(
            popup,
            text="➕ Add Category",
            fg_color="#1F6AA5",
            hover_color="#144870",
            font=("Serif Sans", 12),
            command=add_category
        ).grid(row=1, column=0, columnspan=2, pady=10)

        popup.grid_rowconfigure(0, weight=1)
        popup.grid_columnconfigure(0, weight=1)

        refresh()
    # ---------------- STATEMENTS ----------------
    def view_statement(self):
        os.startfile(self.transaction_file)

    def show_monthly_statement(self):
        df = pd.read_excel(self.transaction_file)
        df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y")
        df = df[df["Date"].dt.month == datetime.now().month]
        df["Date"] = df["Date"].dt.strftime("%d-%m-%Y")
        fname = "Monthly_Statement.xlsx"
        df.to_excel(fname, index=False)
        os.startfile(fname)

    def show_date_range_statement(self):
        popup = ctk.CTkToplevel(self.root, fg_color="gray14")
        popup.grab_set()
        self.center_popup(popup, 350, 200)

        start = ctk.StringVar()
        end = ctk.StringVar()

        ctk.CTkLabel(popup, text="Start Date (DD-MM-YYYY)").pack()
        ctk.CTkEntry(popup, textvariable=start).pack()
        ctk.CTkLabel(popup, text="End Date (DD-MM-YYYY)").pack()
        ctk.CTkEntry(popup, textvariable=end).pack()

        def submit():
            s = datetime.strptime(start.get(), "%d-%m-%Y")
            e = datetime.strptime(end.get(), "%d-%m-%Y")
            df = pd.read_excel(self.transaction_file)
            df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y")
            df = df[(df["Date"] >= s) & (df["Date"] <= e)]
            df["Date"] = df["Date"].dt.strftime("%d-%m-%Y")
            fname = f"Statement_{start.get()}_to_{end.get()}.xlsx"
            df.to_excel(fname, index=False)
            os.startfile(fname)
            popup.destroy()

        ctk.CTkButton(popup, text="Generate", command=submit).pack(pady=15)

    # ---------------- ANALYSIS ----------------
    def show_analysis(self):
        popup = ctk.CTkToplevel(self.root, fg_color="gray14")
        popup.grab_set()
        self.center_popup(popup, 350, 200)

        s = ctk.StringVar()
        e = ctk.StringVar()

        ctk.CTkLabel(popup, text="Start Date (DD-MM-YYYY)").pack()
        ctk.CTkEntry(popup, textvariable=s).pack()
        ctk.CTkLabel(popup, text="End Date (DD-MM-YYYY)").pack()
        ctk.CTkEntry(popup, textvariable=e).pack()

        def submit():
            start = datetime.strptime(s.get(), "%d-%m-%Y")
            end = datetime.strptime(e.get(), "%d-%m-%Y")
            df = pd.read_excel(self.transaction_file)
            df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y")
            df = df[(df["Date"] >= start) & (df["Date"] <= end)]

            add = df[df["Type"] == "Add"].groupby("Category")["Amount"].sum()
            rem = df[df["Type"] == "Remove"].groupby("Category")["Amount"].sum()

            self.create_pie(add, rem)
            popup.destroy()

        ctk.CTkButton(popup, text="Analyze", command=submit).pack(pady=15)

    def create_pie(self, add, rem):
        win = ctk.CTkToplevel(self.root, fg_color="gray14")
        self.center_popup(win, 900, 600)

        fig, ax = plt.subplots(1, 2, figsize=(10, 5), facecolor="#1f2933")

        ax[0].pie(add, labels=[f"{k} ₹{v}" for k, v in add.items()],textprops={"color":"white"})
        ax[0].set_title("Added",color="white")

        ax[1].pie(rem, labels=[f"{k} ₹{v}" for k, v in rem.items()],textprops={"color":"white"})
        ax[1].set_title("Removed",color="white")

        canvas = FigureCanvasTkAgg(fig, win)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

    def refresh_cash_from_excel(self):
        self.update_available_cash()
        #messagebox.showinfo("Refreshed", "Cash updated from Excel file")


# ---------------- RUN ----------------
if __name__ == "__main__":
    root = ctk.CTk()
    app = CashManagerApp(root)
    root.mainloop()
